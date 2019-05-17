#!/usr/bin/python3

## excel to csv by Timothy Middelkoop Copyright 2018, Apache License 2.0

import openpyxl
import sys

wb = openpyxl.load_workbook(filename=sys.argv[1], read_only=True, data_only=True)

for sheet in wb.sheetnames:
    ws=wb[sheet]
    print(sheet)

    header=[]
    for cell in ws[1]:
        header.append(str(cell.value))
    print(','.join(header))

    for row in ws.iter_rows(min_row=2):
        line=[]
        for cell in row:
            if(cell.value)==None:
                line.append('')
            elif(cell.data_type==openpyxl.cell.cell.TYPE_NUMERIC):
                line.append(str(cell.value))
            else:
                line.append("'%s'" % cell.value)
        print(','.join(line))

